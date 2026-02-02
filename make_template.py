import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from datetime import date, timedelta

# Output filename
OUTPUT_FILE = "control_tower_input_with_help.xlsx"

# ---------------------------------------------------------
# 1. Define Schematic Data
# ---------------------------------------------------------

# Horizon setup
START_DATE = date(2026, 1, 19)  # Monday
WEEKS = 8

# Master Data
master_data = [
    {"sku": "PARACET500_TAB", "sku_desc": "Paracetamol 500mg", "product_family": "Analgesics", "uom": "EA", "location": "BHI_DC1", "location_type": "DC", "unit_revenue": 12.5, "unit_cogs": 7.1, "holding_cost_rate_pa": 0.18, "shelf_life_days": 730},
    {"sku": "AMOX500_CAP", "sku_desc": "Amoxicillin 500mg", "product_family": "Antibiotics", "uom": "EA", "location": "BHI_DC1", "location_type": "DC", "unit_revenue": 18.0, "unit_cogs": 10.5, "holding_cost_rate_pa": 0.18, "shelf_life_days": 500},
    {"sku": "IBUP400_TAB", "sku_desc": "Ibuprofen 400mg", "product_family": "Analgesics", "uom": "EA", "location": "BHI_DC1", "location_type": "DC", "unit_revenue": 8.0, "unit_cogs": 4.5, "holding_cost_rate_pa": 0.18, "shelf_life_days": 730},
    {"sku": "VITC1000_TAB", "sku_desc": "Vitamin C 1000mg", "product_family": "Vitamins", "uom": "EA", "location": "BHI_DC1", "location_type": "DC", "unit_revenue": 5.0, "unit_cogs": 2.0, "holding_cost_rate_pa": 0.20, "shelf_life_days": 365},
    {"sku": "METFOR500_TAB", "sku_desc": "Metformin 500mg", "product_family": "Antidiabetic", "uom": "EA", "location": "BHI_DC1", "location_type": "DC", "unit_revenue": 3.5, "unit_cogs": 1.2, "holding_cost_rate_pa": 0.15, "shelf_life_days": 1000},
]

# Inventory Sanpshot (as of START_DATE)
inventory_data = [
    {"as_of_date": START_DATE, "sku": "PARACET500_TAB", "location": "BHI_DC1", "on_hand_qty": 450000, "qa_hold_qty": 20000, "blocked_qty": 0, "safety_stock_qty": 120000},
    {"as_of_date": START_DATE, "sku": "AMOX500_CAP", "location": "BHI_DC1", "on_hand_qty": 20000, "qa_hold_qty": 0, "blocked_qty": 0, "safety_stock_qty": 50000},
    {"as_of_date": START_DATE, "sku": "IBUP400_TAB", "location": "BHI_DC1", "on_hand_qty": 800000, "qa_hold_qty": 0, "blocked_qty": 0, "safety_stock_qty": 60000},
    {"as_of_date": START_DATE, "sku": "VITC1000_TAB", "location": "BHI_DC1", "on_hand_qty": 30000, "qa_hold_qty": 5000, "blocked_qty": 0, "safety_stock_qty": 35000},
    {"as_of_date": START_DATE, "sku": "METFOR500_TAB", "location": "BHI_DC1", "on_hand_qty": 150000, "qa_hold_qty": 0, "blocked_qty": 0, "safety_stock_qty": 40000},
]

# Demand Plan (8 weeks)
demand_data = []
skus = [m["sku"] for m in master_data]
base_demands = {"PARACET500_TAB": 80000, "AMOX500_CAP": 40000, "IBUP400_TAB": 20000, "VITC1000_TAB": 15000, "METFOR500_TAB": 30000}

for w in range(WEEKS):
    ws = START_DATE + timedelta(days=7 * w)
    for sku in skus:
        qty = base_demands[sku]
        if w == 4 and sku == "PARACET500_TAB": qty *= 1.5
        demand_data.append({"week_start": ws, "sku": sku, "location": "BHI_DC1", "forecast_qty": qty, "customer_priority": "TRADE"})

# Supply Plan
supply_data = []
supply_data.append({"week_start": START_DATE + timedelta(days=21), "sku": "PARACET500_TAB", "location": "BHI_DC1", "supply_qty": 100000, "supply_source": "PLANT_A", "supply_type": "PROD"})
supply_data.append({"week_start": START_DATE + timedelta(days=42), "sku": "PARACET500_TAB", "location": "BHI_DC1", "supply_qty": 100000, "supply_source": "PLANT_A", "supply_type": "PROD"})
supply_data.append({"week_start": START_DATE + timedelta(days=28), "sku": "AMOX500_CAP", "location": "BHI_DC1", "supply_qty": 150000, "supply_source": "PLANT_B", "supply_type": "PO"})
supply_data.append({"week_start": START_DATE + timedelta(days=7), "sku": "VITC1000_TAB", "location": "BHI_DC1", "supply_qty": 20000, "supply_source": "PLANT_C", "supply_type": "PROD"})
supply_data.append({"week_start": START_DATE + timedelta(days=14), "sku": "METFOR500_TAB", "location": "BHI_DC1", "supply_qty": 50000, "supply_source": "PLANT_A", "supply_type": "PROD"})

# Logistics Lanes
lanes_data = [
    {"from_location": "PLANT_A", "to_location": "BHI_DC1", "mode": "ROAD", "transit_days": 2, "cost_per_unit": 0.05},
    {"from_location": "PLANT_B", "to_location": "BHI_DC1", "mode": "ROAD", "transit_days": 4, "cost_per_unit": 0.08},
]

# Params
params_data = [
    {"param_name": "horizon_weeks", "param_value": 8, "notes": "Planning Horizon"},
    {"param_name": "service_level_target", "param_value": 0.95, "notes": "Target OTIF"},
    {"param_name": "excess_weeks_threshold", "param_value": 12, "notes": "WOC > 12 = Excess"},
]

# Calendar
calendar_data = []
for w in range(WEEKS):
    ws = START_DATE + timedelta(days=7 * w)
    calendar_data.append({"week_start": ws, "week_label": f"{ws.year}-W{ws.isocalendar()[1]:02d}"})

# Help Sheet Data
help_data = [
    {"Section": "Overview", "Instructions": "This tool calculates Supply Chain risks (Stockouts, Revenue at Risk) based on your inputs."},
    {"Section": "Sheet: Inventory", "Instructions": "Snapshot of stock per SKU/Location. 'as_of_date' is the snapshot date."},
    {"Section": "Sheet: Demand_Plan", "Instructions": "Weekly forecast quantities. 'week_start' must be Mondays."},
    {"Section": "Sheet: Supply_Plan", "Instructions": "Incoming supply/production. 'week_start' must be Mondays."},
    {"Section": "Sheet: Master_Data", "Instructions": "Optional. Unit prices (Revenue/COGS) for economic calculations."},
    {"Section": "Important", "Instructions": "Do not rename sheets or columns. Ensure dates are strictly YYYY-MM-DD."},
    {"Section": "Scenarios", "Instructions": "Use the sidebar in the app to simulate Demand Uplift or Supply Delays."}
]

dfs = {
    "Help": pd.DataFrame(help_data),
    "Master_Data": pd.DataFrame(master_data),
    "Inventory": pd.DataFrame(inventory_data),
    "Demand_Plan": pd.DataFrame(demand_data),
    "Supply_Plan": pd.DataFrame(supply_data),
    "Logistics_Lanes": pd.DataFrame(lanes_data),
    "Constraints_Params": pd.DataFrame(params_data),
    "Calendar": pd.DataFrame(calendar_data)
}

# ---------------------------------------------------------
# 3. Write to Excel with formatting
# ---------------------------------------------------------
wb = Workbook()
# Remove default sheet
del wb["Sheet"]

for sheet_name, df in dfs.items():
    ws = wb.create_sheet(sheet_name)
    
    # Write dataframe
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Format Header
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # Auto-filter
    ws.auto_filter.ref = ws.dimensions
    
    # Simple column adjusting
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

wb.save(OUTPUT_FILE)
print(f"✅ Generated {OUTPUT_FILE} successfully.")
